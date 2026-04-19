import json
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO
from http.server import BaseHTTPRequestHandler

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        # Read the JSON sent from React
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        req_data = json.loads(post_data)

        s_date = req_data.get("startDate", "២១")
        e_date = req_data.get("endDate", "០១ ខែមីនា ឆ្នាំ២០២៦")
        r_date = req_data.get("reportDate", "ត្រូវនឹងថ្ងៃទី០១ ខែមីនា ឆ្នាំ២០២៦")
        rows_data = req_data.get("rows", [])

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        font_title_moul = Font(name='Khmer OS Moul Light', size=12)
        font_moul_large = Font(name='Khmer OS Moul Light', size=14)
        font_regular = Font(name='Khmer OS Battambang', size=11)
        font_bold = Font(name='Khmer OS Battambang', size=11, bold=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws['L1'] = "ព្រះរាជាណាចក្រកម្ពុជា"
        ws['L1'].font = font_title_moul
        ws['L1'].alignment = align_center
        ws.merge_cells('L1:P1')

        ws['L2'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
        ws['L2'].font = font_title_moul
        ws['L2'].alignment = align_center
        ws.merge_cells('L2:P2')

        ws['A3'] = "អធិការដ្ឋាននគរបាលស្រុកសាមគ្គីមុនីជ័យ"
        ws['A3'].font = font_title_moul
        ws['A6'] = "លទ្ធផលការងារសង្គមរបស់កងកម្លាំង"
        ws['A6'].font = font_moul_large
        ws['A6'].alignment = align_center
        ws.merge_cells('A6:Q6')
        ws['A7'] = f"ប្រចាំសប្តាហ៍ គិតពីថ្ងៃទី {s_date} ដល់ ថ្ងៃទី​ {e_date}"
        ws['A7'].font = font_regular
        ws['A7'].alignment = align_center
        ws.merge_cells('A7:Q7')

        headers = [
            ("A8", "A10", "ល.រ"), ("B8", "B10", "គោលដៅ\nចុះធ្វើការ"),
            ("C8", "D8", "សម្រួលចរាចរណ៍"), ("C9", "C10", "សាលា\nរៀន"), ("D9", "D10", "ទីប្រ\nជុំជន"),
            ("E8", "H8", "ការផ្តល់សេវារដ្ឋបាល"), ("E9", "E10", "អត្ត.\nខ្មែរ"), ("F9", "F10", "សៀវភៅ\nស្នាក់នៅ"), ("G9", "G10", "សៀវភៅ\nគ្រួសារ"), ("H9", "H10", "ផ្សេងៗ"),
            ("I8", "M8", "អន្តរគមន៍ជួយសង្គ្រោះប្រជាពលរដ្ឋ"), ("I9", "K9", "គ្រោះថ្នាក់"), ("I10", "I10", "ចរាចរណ៍"), ("J10", "J10", "អគ្គីភ័យ"), ("K10", "K10", "អាវុធ.\nជាតិផ្ទុះ"), ("L9", "L10", "អ្នកឆ្លង\nទន្លេ"), ("M9", "M10", "ផ្សេងៗ"),
            ("N8", "Q8", "ការចុះជួយប្រ/រដ្ឋ"), ("N9", "N10", "ជួសជុល\nផ្ទះ"), ("O9", "O10", "ជួយប្រ\nមូលផល"), ("P9", "P10", "ជួយជន\nក្រីក្រ"), ("Q9", "Q10", "ផ្សេងៗ"),
        ]
        for tl, br, txt in headers:
            ws[tl] = txt
            ws[tl].font = font_bold
            ws[tl].alignment = align_center
            if tl != br: ws.merge_cells(f"{tl}:{br}")
        
        for row in ws.iter_rows(min_row=8, max_row=10, min_col=1, max_col=17):
            for cell in row: cell.border = border

        curr_row = 11
        totals = [0] * 15
        for idx, row in enumerate(rows_data):
            ws.cell(row=curr_row, column=1, value=idx+1).border = border
            ws.cell(row=curr_row, column=1).alignment = align_center
            ws.cell(row=curr_row, column=2, value=row[0]).border = border
            
            for c_idx in range(1, 16):
                val = row[c_idx]
                val = int(val) if str(val).isdigit() else 0
                cell = ws.cell(row=curr_row, column=c_idx+2, value=val if val > 0 else "")
                cell.border = border
                cell.alignment = align_center
                totals[c_idx-1] += val
            curr_row += 1

        ws.cell(row=curr_row, column=1, value="សរុប").font = font_bold
        ws.merge_cells(f"A{curr_row}:B{curr_row}")
        ws.cell(row=curr_row, column=1).alignment = align_center
        ws.cell(row=curr_row, column=1).border = border
        ws.cell(row=curr_row, column=2).border = border
        for c_idx, tot in enumerate(totals):
            cell = ws.cell(row=curr_row, column=c_idx+3, value=tot if tot > 0 else "")
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border

        ws.cell(row=curr_row+2, column=11, value="ថ្ងៃអាទិត្យ ១៣កើត ខែផល្គុន ឆ្នាំម្សាញ់ សប្តស័ក ពុទ្ធសករាជ ២៥៦៩").font = font_regular
        ws.cell(row=curr_row+3, column=11, value=r_date).font = font_regular
        ws.cell(row=curr_row+4, column=11, value="នាយប៉ុស្តិ៍នគរបាលរដ្ឋបាលក្រាំងចេក").font = font_bold

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for col in range(3, 18): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 9

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send back to Frontend
        self.send_response(200)
        self.send_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename="Police_Report.xlsx"')
        self.end_headers()
        self.wfile.write(output.read())
        return
