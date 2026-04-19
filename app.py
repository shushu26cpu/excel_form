import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO

# --- App Config ---
st.set_page_config(page_title="ប្រព័ន្ធគ្រប់គ្រងរបាយការណ៍", layout="wide")
st.title("ប្រព័ន្ធបញ្ចូលទិន្នន័យរបាយការណ៍កងកម្លាំង")
st.markdown("បញ្ចូលទិន្នន័យនៅទីនេះ រួចទាញយកជា Excel ឬ PDF។")

# --- Inputs for Dates ---
col1, col2, col3 = st.columns(3)
start_date = col1.text_input("ថ្ងៃទីចាប់ផ្តើម (ឧទា. ២១)", "២១")
end_date = col2.text_input("ថ្ងៃទីបញ្ចប់ (ឧទា. ០១ ខែមីនា ឆ្នាំ២០២៦)", "០១ ខែមីនា ឆ្នាំ២០២៦")
report_date = col3.text_input("ថ្ងៃធ្វើរបាយការណ៍", "ត្រូវនឹងថ្ងៃទី០១ ខែមីនា ឆ្នាំ២០២៦")

# --- Default Data Table ---
default_columns = [
    "គោលដៅចុះធ្វើការ", 
    "សម្រួលចរាចរណ៍ - សាលារៀន", "សម្រួលចរាចរណ៍ - ទីប្រជុំជន",
    "សេវារដ្ឋបាល - អត្ត.ខ្មែរ", "សេវារដ្ឋបាល - ស្នាក់នៅ", "សេវារដ្ឋបាល - គ្រួសារ", "សេវារដ្ឋបាល - ផ្សេងៗ",
    "សង្គ្រោះ - ចរាចរណ៍", "សង្គ្រោះ - អគ្គីភ័យ", "សង្គ្រោះ - អាវុធ", "សង្គ្រោះ - ឆ្លងទន្លេ", "សង្គ្រោះ - ផ្សេងៗ",
    "ជួយប្រជាជន - ផ្ទះ", "ជួយប្រជាជន - មូលផល", "ជួយប្រជាជន - ក្រីក្រ", "ជួយប្រជាជន - ផ្សេងៗ"
]

# Initialize default empty row if not in session state
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame(
        [["ប៉ុស្តិ៍នគរបាលរដ្ឋបាលក្រាំងចេក", 5, 2, 0, 0, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]], 
        columns=default_columns
    )

st.subheader("តារាងទិន្នន័យ (បញ្ចូលទិន្នន័យនៅទីនេះ)")
# Dynamic Data Editor - users can add/delete rows easily
edited_df = st.data_editor(st.session_state.df, num_rows="dynamic", use_container_width=True)

# --- Function to Generate Excel ---
def generate_excel(df, s_date, e_date, r_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    font_title_moul = Font(name='Khmer OS Moul Light', size=12)
    font_moul_large = Font(name='Khmer OS Moul Light', size=14)
    font_regular = Font(name='Khmer OS Battambang', size=11)
    font_bold = Font(name='Khmer OS Battambang', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header Titles
    ws['L1'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['L1'].font = font_title_moul
    ws['L1'].alignment = align_center
    ws.merge_cells('L1:P1')

    ws['L2'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['L2'].font = font_title_moul
    ws['L2'].alignment = align_center
    ws.merge_cells('L2:P2')

    ws['A3'] = "អធិការដ្ឋាននគរបាលស្រុកសាមគ្គីមុនីជ័យ"
    ws['A3'].font = font_title_moul
    ws['A6'] = "លទ្ធផលការងារសង្គមរបស់កងកម្លាំង"
    ws['A6'].font = font_moul_large
    ws['A6'].alignment = align_center
    ws.merge_cells('A6:Q6')
    ws['A7'] = f"ប្រចាំសប្តាហ៍ គិតពីថ្ងៃទី {s_date} ដល់ ថ្ងៃទី​ {e_date}"
    ws['A7'].font = font_regular
    ws['A7'].alignment = align_center
    ws.merge_cells('A7:Q7')

    # Table Headers
    headers = [
        ("A8", "A10", "ល.រ"), ("B8", "B10", "គោលដៅ\nចុះធ្វើការ"),
        ("C8", "D8", "សម្រួលចរាចរណ៍"), ("C9", "C10", "សាលា\nរៀន"), ("D9", "D10", "ទីប្រ\nជុំជន"),
        ("E8", "H8", "ការផ្តល់សេវារដ្ឋបាល"), ("E9", "E10", "អត្ត.\nខ្មែរ"), ("F9", "F10", "សៀវភៅ\nស្នាក់នៅ"), ("G9", "G10", "សៀវភៅ\nគ្រួសារ"), ("H9", "H10", "ផ្សេងៗ"),
        ("I8", "M8", "អន្តរគមន៍ជួយសង្គ្រោះប្រជាពលរដ្ឋ"), ("I9", "K9", "គ្រោះថ្នាក់"), ("I10", "I10", "ចរាចរណ៍"), ("J10", "J10", "អគ្គីភ័យ"), ("K10", "K10", "អាវុធ.\nជាតិផ្ទុះ"), ("L9", "L10", "អ្នកឆ្លង\nទន្លេ"), ("M9", "M10", "ផ្សេងៗ"),
        ("N8", "Q8", "ការចុះជួយប្រ/រដ្ឋ"), ("N9", "N10", "ជួសជុល\nផ្ទះ"), ("O9", "O10", "ជួយប្រ\nមូលផល"), ("P9", "P10", "ជួយជន\nក្រីក្រ"), ("Q9", "Q10", "ផ្សេងៗ"),
    ]
    for tl, br, txt in headers:
        ws[tl] = txt
        ws[tl].font = font_bold
        ws[tl].alignment = align_center
        if tl != br: ws.merge_cells(f"{tl}:{br}")
    
    for row in ws.iter_rows(min_row=8, max_row=10, min_col=1, max_col=17):
        for cell in row: cell.border = border

    # Data Rows
    curr_row = 11
    totals = [0] * 15
    for idx, row in df.iterrows():
        ws.cell(row=curr_row, column=1, value=idx+1).border = border
        ws.cell(row=curr_row, column=1).alignment = align_center
        ws.cell(row=curr_row, column=2, value=row[0]).border = border
        
        for c_idx in range(1, 16):
            val = row[c_idx] if pd.notna(row[c_idx]) else 0
            val = int(val) if str(val).isdigit() else 0
            cell = ws.cell(row=curr_row, column=c_idx+2, value=val if val > 0 else "")
            cell.border = border
            cell.alignment = align_center
            totals[c_idx-1] += val
        curr_row += 1

    # Totals Row
    ws.cell(row=curr_row, column=1, value="សរុប").font = font_bold
    ws.merge_cells(f"A{curr_row}:B{curr_row}")
    ws.cell(row=curr_row, column=1).alignment = align_center
    ws.cell(row=curr_row, column=1).border = border
    ws.cell(row=curr_row, column=2).border = border
    for c_idx, tot in enumerate(totals):
        cell = ws.cell(row=curr_row, column=c_idx+3, value=tot if tot > 0 else "")
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border

    # Footer
    ws.cell(row=curr_row+2, column=11, value="ថ្ងៃអាទិត្យ ១៣កើត ខែផល្គុន ឆ្នាំម្សាញ់ សប្តស័ក ពុទ្ធសករាជ ២៥៦៩").font = font_regular
    ws.cell(row=curr_row+3, column=11, value=r_date).font = font_regular
    ws.cell(row=curr_row+4, column=11, value="នាយប៉ុស្តិ៍នគរបាលរដ្ឋបាលក្រាំងចេក").font = font_bold

    # Styling
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    for col in range(3, 18): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 9

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- Generate Download Buttons ---
st.divider()
st.subheader("ទាញយករបាយការណ៍ (Download Reports)")
col_a, col_b = st.columns(2)

# 1. Excel Export
excel_data = generate_excel(edited_df, start_date, end_date, report_date)
col_a.download_button(
    label="📥 ទាញយកជា Excel (Download Excel)",
    data=excel_data,
    file_name="Police_Report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

# 2. PDF / HTML Export Route
# We generate a formatted HTML document. The user clicks download, opens it, and hits Ctrl+P to save as perfectly styled PDF.
html_content = f"""
<!DOCTYPE html>
<html lang="km">
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Khmer OS Battambang', Arial, sans-serif; font-size: 14px; }}
        .moul {{ font-family: 'Khmer OS Moul Light', cursive; }}
        .center {{ text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ border: 1px solid black; padding: 5px; text-align: center; }}
        .no-border {{ border: none; }}
        .right-text {{ text-align: right; padding-right: 50px; }}
    </style>
</head>
<body>
    <div style="float: right; text-align: center;" class="moul">
        ព្រះរាជាណាចក្រកម្ពុជា<br>ជាតិ សាសនា ព្រះមហាក្សត្រ<br>
    </div>
    <div style="clear: both;"></div>
    <div class="moul" style="margin-top: -30px;">អធិការដ្ឋាននគរបាលស្រុកសាមគ្គីមុនីជ័យ</div>
    
    <h3 class="moul center" style="margin-top: 40px;">លទ្ធផលការងារសង្គមរបស់កងកម្លាំង</h3>
    <div class="center">ប្រចាំសប្តាហ៍ គិតពីថ្ងៃទី {start_date} ដល់ ថ្ងៃទី​ {end_date}</div>

    <table>
        <tr style="font-weight: bold; background-color: #f2f2f2;">
            <td rowspan="3">ល.រ</td><td rowspan="3">គោលដៅចុះធ្វើការ</td><td colspan="2">សម្រួលចរាចរណ៍</td>
            <td colspan="4">ការផ្តល់សេវារដ្ឋបាល</td><td colspan="5">អន្តរគមន៍ជួយសង្គ្រោះប្រជាពលរដ្ឋ</td>
            <td colspan="4">ការចុះជួយប្រ/រដ្ឋ</td>
        </tr>
        <tr style="font-weight: bold; background-color: #f2f2f2;">
            <td rowspan="2">សាលារៀន</td><td rowspan="2">ទីប្រជុំជន</td>
            <td rowspan="2">អត្ត.<br>ខ្មែរ</td><td rowspan="2">សៀវភៅ<br>ស្នាក់នៅ</td><td rowspan="2">សៀវភៅ<br>គ្រួសារ</td><td rowspan="2">ផ្សេងៗ</td>
            <td colspan="3">គ្រោះថ្នាក់</td><td rowspan="2">អ្នកឆ្លង<br>ទន្លេ</td><td rowspan="2">ផ្សេងៗ</td>
            <td rowspan="2">ជួសជុល<br>ផ្ទះ</td><td rowspan="2">ជួយប្រ<br>មូលផល</td><td rowspan="2">ជួយជន<br>ក្រីក្រ</td><td rowspan="2">ផ្សេងៗ</td>
        </tr>
        <tr style="font-weight: bold; background-color: #f2f2f2;">
            <td>ចរាចរណ៍</td><td>អគ្គីភ័យ</td><td>អាវុធ.<br>ជាតិផ្ទុះ</td>
        </tr>
"""
totals = [0]*15
for idx, row in edited_df.iterrows():
    html_content += f"<tr><td>{idx+1}</td><td style='text-align:left;'>{row[0]}</td>"
    for c in range(1, 16):
        val = row[c] if pd.notna(row[c]) and str(row[c]).isdigit() else 0
        html_content += f"<td>{val if int(val)>0 else ''}</td>"
        totals[c-1] += int(val)
    html_content += "</tr>"

html_content += "<tr><td colspan='2' style='font-weight:bold;'>សរុប</td>"
for t in totals: html_content += f"<td style='font-weight:bold;'>{t if t>0 else ''}</td>"
html_content += f"""
    </table>
    <div class="right-text" style="margin-top: 30px;">
        ថ្ងៃអាទិត្យ ១៣កើត ខែផល្គុន ឆ្នាំម្សាញ់ សប្តស័ក ពុទ្ធសករាជ ២៥៦៩<br>
        {report_date}<br><br>
        <b class="moul">នាយប៉ុស្តិ៍នគរបាលរដ្ឋបាលក្រាំងចេក</b>
    </div>
</body></html>
"""

col_b.download_button(
    label="📄 ទាញយកជា PDF (Print to PDF)",
    data=html_content,
    file_name="Police_Report.html",
    mime="text/html",
    use_container_width=True
)

st.info("💡 **របៀប Save ជា PDF:** សូមចុចប៊ូតុង **ទាញយកជា PDF** ខាងលើ វាស្រួលជាងគេ។ វានឹងទាញយក File HTML មួយ។ សូមបើក File នោះនៅលើ Browser រួចចុច `Ctrl + P` (ឬ ចុច Print) រួចជ្រើសរើសយកពាក្យ **Save as PDF**។ វិធីនេះធានាថា អក្សរខ្មែរនឹងមិនខូចទ្រង់ទ្រាយឡើយ។")
